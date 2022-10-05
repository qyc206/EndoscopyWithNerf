'''

convert2npy.py

Converts camera pose information in excel files provided in EndoSLAM dataset 
(https://github.com/CapsuleEndoscope/EndoSLAM) into the format expected by 
NeRF model for training. The expected format is described in the following 
link: https://github.com/Fyusion/LLFF/issues/10.

Flags:
    --xlsx (Path to the excel file with camera pose information)
    --output (Path to the directory where the produced poses_bounds.npy file can be placed)
    --npy (Path to the poses_bounds.npy file, use for viewing the contents of the file)

Additional Notes:
    This code is currently set to read and convert only the first 20 camera poses from 
    the input excel file. Please modify the convert_to_npy function accordingly if you 
    would like to include all or a different number of camera poses for your conversion.

Feel free to reach out via email with any questions/concerns: qyc206@nyu.edu

'''

import argparse
import sys
import os

import numpy as np
import openpyxl as opx
import csv

from scipy.spatial.transform import Rotation as R

# arbitrary values
# IMAGE_VEC = np.array([[480, 640, 28]])  # [image height, image width, focal length (mm)]
IMAGE_VEC = np.array([[320, 320, 28]])  # [image height, image width, focal length (mm)]
CLOSE_DEPTH = 0.75
FAR_DEPTH = 2

# CLOSE_DEPTH = 0.01
# FAR_DEPTH = 2

def read_npy_file(filePath):
    try: 
        data = np.load(filePath)
        return data
    except (IOError, ValueError) as e:
        print(e)
        sys.exit(1)

def convert_quot2rotMatrix(quot):
    # q = w + xi + yj + zk or q = q0 + q1i + q2j + q3k
    q0 = quot[3] # w
    q1 = quot[0] # x
    q2 = quot[1] # y
    q3 = quot[2] # z

    # R(x y z) = (r1 r2 r3) 
    r1 = np.array([[(2*(q0*q0 + q1*q1))-1], [2*(q1*q2 + q0*q3)], [2*(q1*q3 - q0*q2)]])
    r2 = np.array([[2*(q1*q2 - q0*q3)], [(2*(q0*q0 + q2*q2))-1], [2*(q2*q3 + q0*q1)]])
    r3 = np.array([[2*(q1*q3 + q0*q2)], [2*(q2*q3 - q0*q1)], [(2*(q0*q0 + q3*q3))-1]])

    # change columns to [down, right, backwards] or [-y, x, z]
    r2 = -1*r2
    return np.concatenate((r2, r1, r3), axis=1)

def convert_quot2rotMatrix_usingScipy(quot):
    # q = w + xi + yj + zk or q = q0 + q1i + q2j + q3k 
    x = quot[0] 
    y = quot[1] 
    z = quot[2] 
    w = quot[3]

    return R.from_quat([x, y, z, w]).as_matrix()

def convert_to_npy_csv(csvPath, npyPath):
    # explicitly state number of images desired
    numImages = 20

    # initialize an empty Nx17 numpy array
    posesBoundsMatrix = np.empty((numImages,17))   # only the first numImages images

    with open(csvPath) as csv_file:
        csv_reader = csv.reader(csv_file, delimiter=',')
        curr_row = 0
        for row in csv_reader:
            if (curr_row == 0):
                curr_row += 1
                continue

            if (curr_row-1 == numImages):
                break

            # get translation matrix
            tranMatrix = np.array([[float(row[0]), float(row[1]), float(row[2])]])

            # get rotation matrix
            rotMatrix = convert_quot2rotMatrix_usingScipy((float(row[3]), float(row[4]), float(row[5]), float(row[6])))

            # concatenate & flatten rotation, translation and [height, width, focal] matrices
            poseMatrixFlat = (np.concatenate((rotMatrix, tranMatrix.T, IMAGE_VEC.T), axis=1)).flatten()
            # concatenate close/far depths to get 17d array for pose & add to poses matrix
            finalMatrix = np.concatenate((poseMatrixFlat, CLOSE_DEPTH, FAR_DEPTH), axis=None)
            posesBoundsMatrix[curr_row-1] = finalMatrix

            curr_row += 1
        
        # write Nx17 numpy array to .npy file
        with open(npyPath+"/poses_bounds.npy", 'wb') as writeFile:
            np.save(writeFile, posesBoundsMatrix)

def convert_to_npy(xlsxPath, npyPath):
    # read excel file
    workbook = opx.load_workbook(xlsxPath)
    # read active sheet from excel file
    sheet = workbook.active

    # explicitly state number of images desired
    numImages = 20
    # numImages = 120

    # initialize an empty Nx17 numpy array
    # posesBoundsMatrix = np.empty((sheet.max_row-1,17))    # use this instead for ALL images in excel file
    posesBoundsMatrix = np.empty((numImages,17))   # only the first numImages images
    currRow = 0

    # iterate through images and fill poses matrix
    # for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row): # use this instead for ALL images in excel file
    for row in sheet.iter_rows(min_row=2, max_row=numImages+1): # only the first numImages images
        # get translation matrix
        tranMatrix = np.array([[row[3].value, row[4].value, row[5].value]])

        # get rotation matrix
        # rotMatrix = convert_quot2rotMatrix((row[6].value, row[7].value, row[8].value, row[9].value))
        rotMatrix = convert_quot2rotMatrix_usingScipy((row[6].value, row[7].value, row[8].value, row[9].value))

        # concatenate & flatten rotation, translation and [height, width, focal] matrices
        poseMatrixFlat = (np.concatenate((rotMatrix, tranMatrix.T, IMAGE_VEC.T), axis=1)).flatten()
        # concatenate close/far depths to get 17d array for pose & add to poses matrix
        finalMatrix = np.concatenate((poseMatrixFlat, CLOSE_DEPTH, FAR_DEPTH), axis=None)
        posesBoundsMatrix[currRow] = finalMatrix
        currRow = currRow + 1

    # close excel file
    workbook.close()
    
    # write Nx17 numpy array to .npy file
    with open(npyPath+"/poses_bounds.npy", 'wb') as writeFile:
        np.save(writeFile, posesBoundsMatrix)

def main(args):
    if (args.npy):
        data = read_npy_file(args.npy)
        print(data)
        print(len(data))
    elif (args.xlsx):
        convert_to_npy(args.xlsx, args.output)
    elif (args.csv):
        convert_to_npy_csv(args.csv, args.output)
    else:
        parser.print_help(sys.stderr)
        sys.exit(1)

if __name__=='__main__':
    parser = argparse.ArgumentParser()

    parser.add_argument('--xlsx', type=str, help='Path to poses excel file')
    parser.add_argument('--csv', type=str, help='Path to poses csv file')
    parser.add_argument('--output', type=str, default=os.getcwd(), help='Path to directory to place the produced poses_bounds.npy file')
    parser.add_argument('--npy', type=str, help='Path to poses_bounds.npy file')
    
    args = parser.parse_args()  # retrieve arguments

    main(args)