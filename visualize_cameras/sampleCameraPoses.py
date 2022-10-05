import os
import openpyxl as opx
from scipy.spatial.transform import Rotation as R


XLSX_PATH = '/Downloads/nerf_related/HighCam/Stomach-III/TumorfreeTrajectory_4/Poses/low_high_pose_stom3_teste4_high_images.xlsx'
SAVE_PATH = '/Downloads/nerf_related'

CAMERA_INFO = []
STEPS = 25


def readXlsxInfo2Lst():
    workbook = opx.load_workbook(XLSX_PATH)
    sheet = workbook.active

    count = 0

    for row in sheet.iter_rows(min_row=2):
        if row[0].value%STEPS == 0 and count < 20:
            CAMERA_INFO.append([cell.value for cell in row])
            count += 1

            if row[0].value+STEPS > sheet.max_row:
                return False

    workbook.close()
    return True

def updateCamera(deltaDeg, deltaZ):
    for i in range(len(CAMERA_INFO)):
        cameraInfo = CAMERA_INFO[i]
        rEuler = R.from_quat([cameraInfo[6], cameraInfo[7], cameraInfo[8], cameraInfo[9]]).as_euler('xyz', degrees=True)
        rFlipped = R.from_euler('xyz', [rEuler[0], rEuler[1]+deltaDeg, rEuler[2]], degrees=True).as_quat()

        newCameraInfo = CAMERA_INFO[i][:5]
        newCameraInfo.extend([CAMERA_INFO[i][5] + deltaZ])
        newCameraInfo.extend(rFlipped)
        CAMERA_INFO[i] = newCameraInfo

def saveCameraInfo2Excel():
    workbook = opx.Workbook()
    sheet = workbook.active

    sheet.append(['', 'ImageFrame', 'Pose_Index', 'trans_x', 'trans_y', 'trans_z', 'quot_x', 'quot_y', 'quot_z', 'quot_w'])

    for row in CAMERA_INFO:
        sheet.append(row)

    workbook.save(SAVE_PATH+"/sampled_poses_step" + str(STEPS) + ".xlsx")


success = readXlsxInfo2Lst()
if success:
    updateCamera(180, 37)
    saveCameraInfo2Excel()
else:
    print("not enough rows... try a smaller STEP value")