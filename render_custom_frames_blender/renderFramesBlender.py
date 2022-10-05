## NOTE: parts of this code are inspired by Yujie's codes

import os
import openpyxl as opx
from scipy.spatial.transform import Rotation as R

import json
import bpy
import math
import numpy as np

'''

SET UP PORTION

'''

# provide an excel file with camera info
XLSX_PATH = '/Downloads/nerf_related/results/custom_frames_camera_info.xlsx'
CONVERT2LST = True
CAMERA_INFO = []    # if a file is not provided, replace this with camera info

# update camera (i.e. flip upside_down & shift up via z-axis)
UPDATE_CAMERA = False

# FRAMES = 20
FRAMES = 2

resolution_x = 640
resolution_y = 480
DEPTH_SCALE = 1.4
COLOR_DEPTH = 8
FORMAT = 'PNG'

RESULTS_PATH = '/Downloads/nerf_related/results'
FRAMES_PATH = '/Downloads/nerf_related/results/frames'

# create excel with camera information
CREATE_EXCEL = False

# render depths
RENDER_DEPTH = True

DEBUG = False

def readXlsxInfo2Lst():
    workbook = opx.load_workbook(XLSX_PATH)
    sheet = workbook.active

    for row in sheet.iter_rows(min_row=2, max_row=FRAMES+1):
        # [trans_x, trans_y, trans_z, quot_x, quot_y, quot_z, quot_w]
        CAMERA_INFO.append([row[3].value, row[4].value, row[5].value, row[6].value, row[7].value, row[8].value, row[9].value])

    workbook.close()

if CONVERT2LST:
    readXlsxInfo2Lst()

def updateCamera(deltaDeg, deltaZ):
    for i in range(len(CAMERA_INFO)):
        cameraInfo = CAMERA_INFO[i]
        rEuler = R.from_quat([cameraInfo[3], cameraInfo[4], cameraInfo[5], cameraInfo[6]]).as_euler('xyz', degrees=True)
        rFlipped = R.from_euler('xyz', [rEuler[0], rEuler[1]+deltaDeg, rEuler[2]], degrees=True).as_quat()

        newCameraInfo = CAMERA_INFO[i][:2]
        newCameraInfo.extend([CAMERA_INFO[i][2] + deltaZ])
        newCameraInfo.extend(rFlipped)
        CAMERA_INFO[i] = newCameraInfo

if UPDATE_CAMERA:
    updateCamera(180, 37)

def saveCameraInfo2Excel():
    workbook = opx.Workbook()
    sheet = workbook.active

    sheet.append(['', 'ImageFrame', 'Pose_Index', 'trans_x', 'trans_y', 'trans_z', 'quot_x', 'quot_y', 'quot_z', 'quot_w'])

    for i in range(len(CAMERA_INFO)):
        row = [i, 0, 0]
        row.extend(CAMERA_INFO[i])
        sheet.append(row)

    workbook.save(RESULTS_PATH+"/custom_frames_camera_info.xlsx")

if CREATE_EXCEL:
    saveCameraInfo2Excel()


'''

RENDERING PORTION

'''

# data to store in JSON file
camera = bpy.data.objects['Camera']
camera.data.lens_unit = 'FOV'
camera.data.angle = math.radians(26.20)
out_data = {
    'camera_angle_x': bpy.data.objects['Camera'].data.angle_x,
}

# render optimizations
bpy.context.scene.render.use_persistent_data = True

# Set up rendering of depth map.
bpy.context.scene.use_nodes = True
tree = bpy.context.scene.node_tree
links = tree.links

# add passes for additionally dumping albedo and normals.
bpy.context.view_layer.use_pass_normal = True
bpy.context.view_layer.cycles.denoising_store_passes = True
bpy.context.scene.render.image_settings.file_format = str(FORMAT)
if RENDER_DEPTH:
    bpy.context.scene.render.image_settings.color_depth = str(COLOR_DEPTH)

if not DEBUG:
    # Create input render layer node.
    render_layers = tree.nodes.new('CompositorNodeRLayers')
    image_file_output = tree.nodes.new(type='CompositorNodeOutputFile')
    image_file_output.label = "Image_Output"
    image_file_output.base_path = RESULTS_PATH

    if RENDER_DEPTH:
        depth_file_output = tree.nodes.new(type="CompositorNodeOutputFile")
        depth_file_output.label = 'Depth Output'
        depth_file_output.base_path = RESULTS_PATH
        depth_file_output.format.file_format = "OPEN_EXR"
        
        map = tree.nodes.new(type="CompositorNodeMapRange")
        map.inputs[1].default_value = 0.0
        map.inputs[2].default_value = 5.0
        map.inputs[3].default_value = 0.0
        map.inputs[4].default_value = 255.0   
        links.new(render_layers.outputs['Depth'], map.inputs[0])
        links.new(map.outputs[0], depth_file_output.inputs[0])

    denoise_node = tree.nodes.new(type='CompositorNodeDenoise')
    links.new(render_layers.outputs['Noisy Image'], denoise_node.inputs[0])
    links.new(render_layers.outputs['Denoising Normal'], denoise_node.inputs[1])
    links.new(render_layers.outputs['Denoising Albedo'], denoise_node.inputs[2])
  
    normal_file_output = tree.nodes.new(type="CompositorNodeOutputFile")
    normal_file_output.label = 'Normal Output'
    links.new(render_layers.outputs['Normal'], normal_file_output.inputs[0])
    links.new(render_layers.outputs['Image'], image_file_output.inputs[0])
    

    viewer_node = tree.nodes.new('CompositorNodeViewer')   
    viewer_node.use_alpha = True
    links.new(denoise_node.outputs['Image'], viewer_node.inputs['Image'])
    links.new(render_layers.outputs['Depth'], viewer_node.inputs['Alpha'])

# Background
bpy.context.scene.render.dither_intensity = 0.0
bpy.context.scene.render.film_transparent = True

# Create collection for objects not to render with background

objs = [ob for ob in bpy.context.scene.objects if ob.type in ('EMPTY') and 'Empty' in ob.name]
bpy.ops.object.delete({"selected_objects": objs})

def parent_obj_to_camera(b_camera):
    origin = (0, 0, 0)
    b_empty = bpy.data.objects.new("Empty", None)
    b_empty.location = origin
    b_camera.parent = b_empty  # setup parenting

    scn = bpy.context.scene
    scn.collection.objects.link(b_empty)
    bpy.context.view_layer.objects.active = b_empty
    # scn.objects.active = b_empty
    return b_empty

scene = bpy.context.scene
scene.render.resolution_x = resolution_x    # width
scene.render.resolution_y = resolution_y    # height
scene.render.resolution_percentage = 100

cam = scene.objects['Camera']
cam_constraint = cam.constraints.new(type='TRACK_TO')
cam_constraint.track_axis = 'TRACK_NEGATIVE_Z'
cam_constraint.up_axis = 'UP_Y'
b_empty = parent_obj_to_camera(cam)  # NOT USED IN FOR LOOP TO RENDER FRAMES
cam_constraint.target = b_empty
cam.rotation_mode = 'QUATERNION'

scene.render.image_settings.file_format = FORMAT  # set output format to .png

if not DEBUG:
    normal_file_output.base_path = ''
    if RENDER_DEPTH:
        depth_file_output.base_path = ''

out_data['frames'] = []

# set up path to store rendered frames
fp = bpy.path.abspath(f"//{FRAMES_PATH}")
if not os.path.exists(fp):
    os.makedirs(fp)

for i in range(0, FRAMES):
    cam.location = (CAMERA_INFO[i][0], CAMERA_INFO[i][1], CAMERA_INFO[i][2])
    cam.rotation_quaternion = (CAMERA_INFO[i][3], CAMERA_INFO[i][4], CAMERA_INFO[i][5], CAMERA_INFO[i][6]) # (w,x,y,z)

    print("Rendering Frame {}...".format(i))
    print("\tRotation in quaternion (x,y,z,w):\n\t{}\n".format(CAMERA_INFO[i][3:]))
    scene.render.filepath = fp + '/r_{0:03d}'.format(i)

    normal_file_output.file_slots[0].path = scene.render.filepath + "_normal77_"
    if RENDER_DEPTH:
        depth_file_output.file_slots[0].path = scene.render.filepath + "_depth77_"

    if DEBUG:
        break
    else:
        bpy.ops.render.render(write_still=True)  # render still

    if RENDER_DEPTH:
        # render original depth values to npy file.
        pixels = np.array(bpy.data.images['Viewer Node'].pixels)
        print(len(pixels))
        width = bpy.context.scene.render.resolution_x 
        height = bpy.context.scene.render.resolution_y

        #reshaping into image array 4 channel (rgbz)
        image = pixels.reshape(height,width,4)

        #depth analysis...
        depth_array = np.asarray(image[:,:,3])
        print(np.max(depth_array), np.min(depth_array))

        print(depth_array.shape)
        #np_depth_path = dir_cur + dir_dataset + str(s_iter).zfill(6)+"_"+ str(f_iter).zfill(2) + 'Depth.npy'
        np_depth_path = scene.render.filepath + '_depth92.npz'
        depth_array = np.array(depth_array[::-1, ...])
        np.savez_compressed(np_depth_path, depth=depth_array)
    
    frame_data = {
        'file_path': scene.render.filepath,
        'rotation (quaternion wxyz)': cam.rotation_quaternion,
        'transform_matrix': listify_matrix(cam.matrix_world)
    }
    out_data['frames'].append(frame_data)
   
print('b_empty.location is ', b_empty.location)

if not DEBUG:
    with open(fp + '/' + 'transforms.json', 'w') as out_file:
        json.dump(out_data, out_file, indent=4)
